#!/usr/bin/python3

import pandas as pd
import os
import sys
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from datetime import datetime
from ipaddress import ip_network, ip_address


# possible arg - report.csv


# Define cell background colors
fill_none = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type=None)
fill_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
fill_red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
fill_yellow = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
fill_orange = PatternFill(start_color="FFD580", end_color="FFD580", fill_type="solid")


#WithSwVersions = True
WithSwVersions = False




# Check if file exists and not zero-sized
def file_exists(pathtofile):
  fileok = False
  if (os.path.isfile(pathtofile)) and (os.path.getsize(pathtofile) != 0):
    fileok = True
  return fileok


# Set the current date
current_date = datetime.now().strftime("%Y%m%d")

# Set the CSV and XLSX file names
default_csv_file = f'report-{current_date}.csv'
csv_file = sys.argv[1] if len(sys.argv) > 1 else default_csv_file
xls_file = f'report-{current_date}.xlsx'


for ff in [csv_file]:
  if not file_exists(ff):
    print(f"The file {ff} isn't exist or empty. Please put a non-empty file in the current directory.")
    exit()



# Read the 1st part of the CSV file (with 2 columns)
df1 = pd.read_csv(csv_file, nrows=1, header=None)

# Read the 2nd part of the CSV file (with 3 columns)
df2 = pd.read_csv(csv_file, skiprows=1, nrows=5, header=None)

# Read the 3rd part of the CSV file (with 6 columns)
df3 = pd.read_csv(csv_file, skiprows=6, header=None)

# Concatenate the two parts
df = pd.concat([df1, df2, df3])

# Create a new workbook and add a worksheet with the current date as the name
wb = Workbook()
ws = wb.active
ws.title = current_date

# Convert the DataFrame to rows and add them to the worksheet
for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)

# Iterate through the rows
for row in ws.iter_rows(min_row=8, max_row=ws.max_row):
    fill = fill_none
    valuation_row=5
    if WithSwVersions:
        valuation_row += 1
    valuation = row[valuation_row].value
    # Check the valuation for fill color:
    if str(valuation).startswith('Green'):
        fill = fill_green
    if str(valuation).startswith('Yellow'):
        fill = fill_yellow
    if str(valuation).startswith('Red'):
        fill = fill_red
    if str(valuation).startswith('Orange'):
        fill = fill_orange
    for cell in row:
        cell.fill = fill


# put comment on Cell Valuation:
comment_valuation = Comment("Valuations:\n\nGreen: device is online, has 2 apps, and is not at fabric\n\nOrange: device is online, not at fabric, but doesn't have 2 apps\n\nYellow: device is online, but still is at fabric\n\nRed: device is offline\n\nRed: provisioned (never onboareded yet)\n", "Sergei")
comment_cell="F8"
if WithSwVersions:
    comment_cell="G8"
ws[comment_cell].comment = comment_valuation

# set cell format as number without signs after decimal dot:
ws['B2'].number_format = '0'


# adjust cells' width:
for column in ws.columns:
    max_length = 0
    column = get_column_letter(column[0].column)  # Get the column name
    for cell in ws[column]:  # Go through all cells in the column
        try: 
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column].width = adjusted_width

# align rights cell C3:C7
for row in range(3, 8):  # Loop over rows 3 to 7 inclusive
    cell = ws['C{}'.format(row)]  # Get the cell at position (row, 3) (column C)
    cell.alignment = Alignment(horizontal='right')  # Set the alignment of the cell to right

# Save the workbook as an XLSX file
wb.save(xls_file)
